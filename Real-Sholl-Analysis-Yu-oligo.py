"""
Sholl Analysis for OPCs

"""

import cv2
import numpy as np
import matplotlib.pyplot as plt
from skimage.morphology import (
    skeletonize,
    remove_small_objects,
    remove_small_holes
)
from skimage.draw import disk
from skimage.draw import line
from skimage.measure import label
from skimage.filters import sato
from skimage.exposure import rescale_intensity
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter



class ZoomPanViewer:
    def __init__(self, image, window_name="viewer", click_callback=None, drag_threshold=5):  # **CHANGED** added drag_threshold
        """
        image: original image (BGR or gray)
        click_callback: function(x, y) in ORIGINAL image coordinates
        """
        self.img = image
        self.win = window_name
        self.click_callback = click_callback

        self.h, self.w = image.shape[:2]

        # View state
        self.scale = 1.0
        self.min_scale = 0.05
        self.max_scale = 50.0
        self.tx = 0.0
        self.ty = 0.0

        # Mouse state
        self.dragging = False
        self.last_x = 0
        self.last_y = 0
        self.down_x = 0              # **ADDED**
        self.down_y = 0              # **ADDED**
        self.moved = False           # **ADDED**
        self.drag_threshold = drag_threshold  # **ADDED**

        cv2.namedWindow(self.win, cv2.WINDOW_NORMAL)
        cv2.setMouseCallback(self.win, self.on_mouse)

    # ---------- Coordinate mapping ----------
    def screen_to_image(self, x, y):
        ix = int((x - self.tx) / self.scale)
        iy = int((y - self.ty) / self.scale)
        ix = max(0, min(self.w - 1, ix))
        iy = max(0, min(self.h - 1, iy))
        return ix, iy

    # ---------- Mouse handler ----------
    def on_mouse(self, event, x, y, flags, param):
        # **CHANGED** Completely rewrote LButton logic to fix duplicate EVENT_LBUTTONDOWN bug
        # Start tracking (potential pan OR click)
        if event == cv2.EVENT_LBUTTONDOWN:                # **CHANGED**
            self.dragging = True                          # **CHANGED**
            self.moved = False                            # **ADDED**
            self.down_x, self.down_y = x, y               # **ADDED**
            self.last_x, self.last_y = x, y               # **CHANGED**

        # Pan while dragging
        elif event == cv2.EVENT_MOUSEMOVE and self.dragging:  # **CHANGED**
            dx = x - self.last_x
            dy = y - self.last_y
            self.tx += dx
            self.ty += dy
            self.last_x, self.last_y = x, y

            if not self.moved:                            # **ADDED**
                if abs(x - self.down_x) + abs(y - self.down_y) >= self.drag_threshold:  # **ADDED**
                    self.moved = True                     # **ADDED**

        # Release: if not moved much, treat as click
        elif event == cv2.EVENT_LBUTTONUP:                # **CHANGED**
            self.dragging = False                         # **CHANGED**
            if not self.moved and self.click_callback is not None:  # **CHANGED**
                ix, iy = self.screen_to_image(x, y)       # **CHANGED**
                self.click_callback(ix, iy)               # **CHANGED**

        # Zoom (unchanged)
        elif event == cv2.EVENT_MOUSEWHEEL:
            zoom_in = flags > 0
            factor = 1.2 if zoom_in else (1 / 1.2)

            old_scale = self.scale
            self.scale = max(self.min_scale,
                             min(self.max_scale, self.scale * factor))

            # zoom around cursor
            ix = (x - self.tx) / old_scale
            iy = (y - self.ty) / old_scale
            self.tx = x - self.scale * ix
            self.ty = y - self.scale * iy

        # Reset (unchanged)
        elif event == cv2.EVENT_LBUTTONDBLCLK:
            self.scale = 1.0
            self.tx = 0.0
            self.ty = 0.0

    # ---------- Render ----------
    def render(self):
        M = np.array([[self.scale, 0, self.tx],
                      [0, self.scale, self.ty]], dtype=np.float32)

        _, _, ww, hh = cv2.getWindowImageRect(self.win)
        if ww <= 0 or hh <= 0:
            hh, ww = 800, 1200

        view = cv2.warpAffine(
            self.img, M, (ww, hh),
            flags=cv2.INTER_NEAREST,
            borderMode=cv2.BORDER_CONSTANT,
            borderValue=0
        )
        return view

    # ---------- Main loop ----------
    def loop(self):
        while True:
            cv2.imshow(self.win, self.render())
            key = cv2.waitKey(16) & 0xFF
            if key in (27, ord('q')):
                break

        cv2.destroyWindow(self.win)

### -----------------------------------------------------------------------


def get_multiple_soma_centers(image):
    """
    Allow manual selection of multiple soma centers via mouse clicks with improved visibility.
    Soma centers are identified by clicking on the image, and they are marked with a visible circle and label.
    """
    soma_coords = []  # List to store coordinates of soma centers
    window_name = "Select Multiple Somas (wheel=zoom, drag=pan, click=mark, q=done)"
    image_copy = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)  # Convert image to color for colored markers
    def on_click(x, y):

        soma_coords.append((x, y))
        # Draw a larger, more visible circle on the image at the clicked position
        cv2.circle(image_copy, (x, y), 10, (0, 255, 0), 2)
        # Add a label next to the circle indicating the order of clicks
        label = str(len(soma_coords))
        cv2.putText(image_copy, label, (x + 15, y + 15), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

    viewer = ZoomPanViewer(  # **ADDED**
            image_copy,  # **ADDED**
            window_name=window_name,  # **ADDED**
            click_callback=on_click  # **ADDED**
        )
    viewer.loop()  # **ADDED**

    cv2.destroyAllWindows()
    return soma_coords

    # Create a window for manual soma selection
    #window_name = "Select Multiple Somas (Click on each soma, press 'q' when done)"
    #image_copy = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)  # Convert image to color for colored markers
    #cv2.namedWindow(window_name)
    # cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    # cv2.setMouseCallback(window_name, mouse_callback)
    # Keep showing the image until 'q' is pressed
    while True:
        cv2.imshow(window_name, image_copy)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cv2.destroyAllWindows()
    return soma_coords


def skeleton_endpoints(skel):
    sk = skel.astype(np.uint8)
    nb = sum(np.roll(np.roll(sk, dy, 0), dx, 1)
             for dy in (-1,0,1) for dx in (-1,0,1)
             if not (dx==0 and dy==0))
    ys, xs = np.where((sk == 1) & (nb == 1))
    return list(zip(xs, ys))

def bridge_endpoints(skel, max_dist=12):
    sk = skel.copy()
    pts = skeleton_endpoints(sk)
    used = set()

    for i, (x1, y1) in enumerate(pts):
        if i in used:
            continue
        for j, (x2, y2) in enumerate(pts):
            if j <= i or j in used:
                continue
            if (x1-x2)**2 + (y1-y2)**2 <= max_dist**2:
                rr, cc = line(y1, x1, y2, x2)
                sk[rr, cc] = True
                used.update([i, j])
                break
    return sk

def keep_component_touching_soma(skel_bool, soma_xy, soma_radius=12):
    """Keep only the skeleton component that touches a small disk around the soma."""
    x, y = soma_xy
    center_rc = (y, x)

    soma_mask = np.zeros_like(skel_bool, dtype=bool)
    rr, cc = disk(center_rc, soma_radius, shape=skel_bool.shape)
    soma_mask[rr, cc] = True

    lab = label(skel_bool, connectivity=2)
    hits = np.unique(lab[soma_mask])
    hits = hits[hits != 0]
    if len(hits) == 0:
        return skel_bool  # nothing touches soma; return as-is for debugging

    keep = hits[0]  # usually 1 component; if multiple, we keep the first
    return lab == keep

def skeleton_length_within_circle(skel_bool, soma_xy, radius):
    """
    Approx skeleton length inside a circle centered at soma_xy.
    Uses 8-neighborhood edge weights: 1 for horizontal/vertical, sqrt(2) for diagonal.
    skel_bool should be a 1-pixel skeleton (bool).
    Returns length in pixels.
    """
    x, y = soma_xy
    center_rc = (y, x)

    h, w = skel_bool.shape
    rr, cc = disk(center_rc, int(radius), shape=(h, w))
    roi = np.zeros((h, w), dtype=bool)
    roi[rr, cc] = True

    s = skel_bool & roi

    # Count edges to right and down to avoid double-counting
    hv = (s[:, :-1] & s[:, 1:]).sum() + (s[:-1, :] & s[1:, :]).sum()
    diag = (s[:-1, :-1] & s[1:, 1:]).sum() + (s[:-1, 1:] & s[1:, :-1]).sum()

    return float(hv + np.sqrt(2) * diag)



def sholl_count_crossings(skeleton_bool, soma_center_xy, radius, thickness=1, max_arc_px=None):
    """
    Count Sholl intersections as number of connected components of (skeleton ∩ ring).
    skeleton_bool: bool image
    soma_center_xy: (x, y) in OpenCV coords
    radius: int
    thickness: ring half-thickness in pixels (1–2 recommended)
    max_arc_px: optional max pixels allowed per component (filters tangential overlaps)
    """
    x, y = soma_center_xy
    center_rc = (y, x)  # **CHANGED** OpenCV (x,y) -> skimage (row,col)

    h, w = skeleton_bool.shape

    outer_r = int(radius + thickness)
    inner_r = int(max(radius - thickness, 0))

    ring = np.zeros((h, w), dtype=bool)
    rr_o, cc_o = disk(center_rc, outer_r, shape=(h, w))
    ring[rr_o, cc_o] = True
    rr_i, cc_i = disk(center_rc, inner_r, shape=(h, w))
    ring[rr_i, cc_i] = False

    inter = skeleton_bool & ring

    lab = label(inter, connectivity=2)
    if lab.max() == 0:
        return 0

    if max_arc_px is None:
        return int(lab.max())

    # Optional: filter very long arcs (tangency)
    count = 0
    for k in range(1, lab.max() + 1):
        if (lab == k).sum() <= max_arc_px:
            count += 1
    return count

def save_sholl_to_excel(xlsx_path, image_path, soma_centers, radii, all_profiles, all_lengths):
    """Append one row per soma to an Excel file."""
    radii_list = [int(r) for r in radii]

    header = ["Image", "Soma_ID", "Soma_X", "Soma_Y", "Skeleton_Length_px"]
    header += [f"Intersections_R{r}px" for r in radii_list]

    # Create folder if needed
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)  # **ADDED**

    # Create or load workbook
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "ShollResults"
        ws.append(header)

        # Style header
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        # Basic column width
        for col in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    image_name = os.path.basename(image_path)

    for idx, ((_, counts), length_px, (x, y)) in enumerate(
        zip(all_profiles, all_lengths, soma_centers), start=1
    ):
        row = [image_name, idx, int(x), int(y), float(length_px)] + [int(c) for c in counts]
        ws.append(row)

    wb.save(xlsx_path)



def process_image_multi_soma(image_path):
    # Load the image in grayscale
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        raise FileNotFoundError(f"Could not read image: {image_path}")

    soma_centers = get_multiple_soma_centers(img)
    if not soma_centers:
        print("No soma selected. Exiting analysis.")
        return None

    imgf = img.astype(np.float32) / 255.0  # **ADDED**
    ridge = sato(imgf, sigmas=range(1, 4), black_ridges=False)  # **ADDED**
    ridge = rescale_intensity(ridge, out_range=(0, 1))  # **ADDED**

    T = np.percentile(ridge, 80)  # **ADDED** (over 80)
    bw_bool = ridge > T  # **CHANGED** boolean mask

    # bridge small gaps in punctate staining
    bw8 = (bw_bool.astype(np.uint8) * 255)  # **CHANGED**
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))  # (try 3,3 or 7,7)
    bw8 = cv2.morphologyEx(bw8, cv2.MORPH_CLOSE, kernel, iterations=1) # use larger iteration for lower quality images
    bw_bool = bw8 > 0

    bw_bool = remove_small_holes(bw_bool, area_threshold=30)  # (tune 30–200)
    skeleton = skeletonize(bw_bool)

    # connect tiny breaks (optional)
    skeleton = bridge_endpoints(skeleton, max_dist=5)
    skeleton = skeleton.astype(bool)  # **CHANGED** make it clean bool once

    # Radii pixel size correspons to 5, 10, 15, 20um to the image processed. Change if your scale is different.
    radii = np.array([31, 62, 92, 123])
    # For quantifying all length wihhin 20um
    R = 123
    all_profiles = []
    img_with_circles = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    all_lengths = []
    skeleton_keep_all = np.zeros_like(skeleton, dtype=bool)  # **ADDED** union of soma-connected skeletons

    for i, soma_center in enumerate(soma_centers):
        skel_one = keep_component_touching_soma(skeleton, soma_center, soma_radius=30)  # **CHANGED** per-soma
        skeleton_keep_all |= skel_one  # **ADDED** accumulate for display


        # ---- LENGTH METRIC ----
        length_123 = skeleton_length_within_circle(
            skel_one, soma_center, R
        )  # **ADDED**

        all_lengths.append(length_123)  # **ADDED**


        print(f"Neuron {i + 1}: skeleton length within R={R}px = {length_123:.2f} px")

        intersection_counts = []
        for radius in radii:
            count = sholl_count_crossings(  # **CHANGED**
                skel_one,  # **CHANGED** use soma-connected skeleton
                soma_center,
                int(radius),
                thickness=1,
                max_arc_px=None
            )
            intersection_counts.append(count)
            cv2.circle(img_with_circles, soma_center, int(radius), (0, 255, 0), 1)

        all_profiles.append((radii, intersection_counts))
        cv2.circle(img_with_circles, soma_center, 10, (255, 0, 0), 2)
        cv2.putText(img_with_circles, f'{i + 1}',
                    (soma_center[0] + 15, soma_center[1] + 15),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)

    # Build ONE skeleton-with-circles image (using the NEW soma-connected skeleton)
    skeleton_with_circles = (skeleton_keep_all.astype(np.uint8) * 255)  # **CHANGED**
    skeleton_with_circles = cv2.cvtColor(skeleton_with_circles, cv2.COLOR_GRAY2BGR)  # **CHANGED**

    for i, soma_center in enumerate(soma_centers):  # **CHANGED**
        for radius in radii:  # **CHANGED**
            cv2.circle(skeleton_with_circles, soma_center, int(radius), (0, 255, 0), 1)  # **CHANGED**
        cv2.circle(skeleton_with_circles, soma_center, 10, (255, 0, 0), 2)  # **CHANGED**
        cv2.putText(skeleton_with_circles, f'{i + 1}',  # **CHANGED**
                    (soma_center[0] + 15, soma_center[1] + 15),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)

    return img, skeleton_keep_all, skeleton_with_circles, img_with_circles, all_profiles, all_lengths, soma_centers, radii



def plot_results(img, skeleton, skeleton_with_circles, img_with_circles, all_profiles):
    """
    Plot the original image, skeletonized image, and image with Sholl analysis circles.
    Additionally, plot the Sholl analysis results for all selected neurons.
    """
    plt.figure(figsize=(30, 13))

    # Plot the original image
    plt.subplot(1, 3, 1)
    plt.imshow(img, cmap='gray')
    plt.title('Original Image')
    plt.axis('off')

    # Plot the skeletonized image with circles
    plt.subplot(1, 3, 2)
    plt.imshow(cv2.cvtColor(skeleton_with_circles, cv2.COLOR_BGR2RGB))
    plt.title('Skeleton + Sholl Circles')
    plt.axis('off')

    # Plot the image with Sholl analysis circles
    plt.subplot(1, 3, 3)
    plt.imshow(cv2.cvtColor(img_with_circles, cv2.COLOR_BGR2RGB))
    plt.title('Image with Sholl Analysis Circles')
    plt.axis('off')

    plt.tight_layout()
    os.makedirs("plots", exist_ok=True)

    base_name = os.path.splitext(os.path.basename(image_path))[0]

    plt.savefig(
        os.path.join("plots", f"{base_name}_plot.png"),
        dpi=300,
        bbox_inches="tight"
    )
    plt.show()


    # Plot Sholl analysis results for all neurons
    # plt.figure(figsize=(10, 6))
    # for i, (radii, counts) in enumerate(all_profiles):
    #     plt.plot(radii, counts, 'o-', label=f'Neuron {i + 1}')
    # plt.title('Sholl Analysis: All Neurons')
    # plt.xlabel('Radius (pixels)')
    # plt.ylabel('Number of Intersections')
    # plt.legend()
    # plt.grid(True)
    # plt.savefig("plot" + f"{image_path[:10]}", dpi=300, bbox_inches="tight")
    # plt.close()
    # plt.show()



def aggregate_profiles(all_profiles):
    """
    Aggregate Sholl profiles across multiple neurons by interpolating them
    to a common set of radii, and then calculating the mean and standard deviation.
    """
    # Find the maximum radius across all profiles
    # max_radius = max(max(profile[0]) for profile in all_profiles)
    # Define a common set of radii for interpolation
    common_radii = np.array([31, 62, 92, 123])

    interpolated_counts = []
    # Interpolate each profile to the common radii
    for radii, counts in all_profiles:
        interp_counts = np.interp(common_radii, radii, counts)
        interpolated_counts.append(interp_counts)

    # Calculate the average and standard deviation across all profiles
    average_profile = np.mean(interpolated_counts, axis=0)
    std_profile = np.std(interpolated_counts, axis=0)

    return common_radii, average_profile, std_profile


def plot_aggregated_results(all_profiles, common_radii, average_profile, std_profile):
    """
    Plot the aggregated Sholl profile, showing the mean and standard deviation,
    along with individual Sholl profiles for each neuron.
    """
    plt.figure(figsize=(10, 6))
    # Plot the aggregated Sholl profile with error bars
    plt.errorbar(common_radii, average_profile, yerr=std_profile, capsize=5, fmt='o-', label='Average Profile')
    plt.fill_between(common_radii, average_profile - std_profile, average_profile + std_profile, alpha=0.3)
    plt.title("Aggregated Sholl Profile")
    plt.xlabel("Radius (pixels)")
    plt.ylabel("Average Number of Intersections")
    plt.legend()
    plt.grid(True)
    plt.show()

    # Plot individual Sholl profiles with the average profile overlaid
    plt.figure(figsize=(10, 6))
    for i, (radii, counts) in enumerate(all_profiles):
        plt.plot(radii, counts, 'o-', alpha=0.7, label=f'Neuron {i + 1}')
    plt.plot(common_radii, average_profile, 'r-', linewidth=2, label='Average Profile')
    plt.title("Individual and Average Sholl Profiles")
    plt.xlabel("Radius (pixels)")
    plt.ylabel("Number of Intersections")
    plt.legend()
    plt.grid(True)
    plt.show()


# Main execution
image_path = "Image-14458.czi-C3.tif"  # Replace with your image path
results = process_image_multi_soma(image_path)

if results:
    img, skeleton, skeleton_with_circles, img_with_circles, all_profiles, all_lengths, soma_centers, radii = results
    # save to excel
    xlsx_path = os.path.join("excel", f"Sholl_{image_path[:11]}.xlsx")  # **ADDED**
    save_sholl_to_excel(xlsx_path, image_path, soma_centers, radii, all_profiles, all_lengths)  # **ADDED**
    print(f"Saved Excel: {xlsx_path}")

    plot_results(img, skeleton,skeleton_with_circles, img_with_circles, all_profiles)
    common_radii, average_profile, std_profile = aggregate_profiles(all_profiles)
    # plot_aggregated_results(all_profiles, common_radii, average_profile, std_profile)
